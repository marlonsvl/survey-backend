from django.contrib import admin
from django.http import HttpResponse
from django.urls import path
from django.shortcuts import render
from django.utils.html import format_html
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from .models import (
    Participant, BergenTikTok, BergenInstagram,
    UCLALoneliness, PrefrontalSymptoms, CAIDS
)


class ParticipantAdmin(admin.ModelAdmin):
    list_display = ['email', 'location', 'age', 'gender', 'university', 'consent_accepted', 'created_at', 'feedback_sent', 'export_button']
    search_fields = ['email', 'country', 'university']
    list_filter = ['location', 'gender', 'marital_status', 'residence_sector', 'socioeconomic_level', 'consent_accepted', 'feedback_sent', 'created_at']
    readonly_fields = ['created_at', 'updated_at', 'consent_date']
    actions = ['export_to_excel_action']
    
    fieldsets = (
        ('Contact Information', {
            'fields': ('email', 'location', 'country')
        }),
        ('Consent', {
            'fields': ('consent_accepted', 'consent_date')
        }),
        ('Personal Information', {
            'fields': ('age', 'gender', 'gender_other', 'marital_status')
        }),
        ('Living Situation', {
            'fields': ('living_with', 'living_with_other')
        }),
        ('Academic Information', {
            'fields': ('university', 'career', 'current_semester', 'gpa_last_semester', 'repeated_cycles', 'repeated_cycles_count')
        }),
        ('Family Information', {
            'fields': ('mother_education_level', 'father_education_level', 'mother_age', 'father_age')
        }),
        ('Socioeconomic Information', {
            'fields': ('residence_sector', 'socioeconomic_level', 'income_sources')
        }),
        ('System Information', {
            'fields': ('created_at', 'updated_at', 'feedback_sent'),
            'classes': ('collapse',)
        }),
    )
    
    def export_button(self, obj):
        return format_html(
            '<a class="button" href="/admin/surveys/export-data/?participant_id={}">Export Data</a>',
            obj.id
        )
    export_button.short_description = 'Actions'
    export_button.allow_tags = True
    
    def export_to_excel_action(self, request, queryset):
        """Admin action to export selected participants to Excel"""
        return self.export_participants_to_excel(queryset)
    
    export_to_excel_action.short_description = "Export selected participants to Excel"
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-data/', self.admin_site.admin_view(self.export_data_view), name='export-data'),
            path('export-all/', self.admin_site.admin_view(self.export_all_view), name='export-all'),
        ]
        return custom_urls + urls
    
    def export_data_view(self, request):
        """View to handle export requests"""
        participant_id = request.GET.get('participant_id')
        
        if participant_id:
            queryset = Participant.objects.filter(id=participant_id)
        else:
            queryset = Participant.objects.all()
        
        return self.export_participants_to_excel(queryset)
    
    def export_all_view(self, request):
        """Export all participants"""
        queryset = Participant.objects.all()
        return self.export_participants_to_excel(queryset)
    
    def export_participants_to_excel(self, queryset):
        """Generate Excel file with all survey data"""
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets for each data type
        self._create_participants_sheet(wb, queryset)
        self._create_bergen_tiktok_sheet(wb, queryset)
        self._create_bergen_instagram_sheet(wb, queryset)
        self._create_ucla_loneliness_sheet(wb, queryset)
        self._create_prefrontal_symptoms_sheet(wb, queryset)
        self._create_caids_sheet(wb, queryset)
        self._create_summary_sheet(wb, queryset)
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename=survey_data_{timestamp}.xlsx'
        
        # Save workbook to response
        wb.save(response)
        return response
    
    def _create_participants_sheet(self, wb, queryset):
        """Create sheet with participant demographic data"""
        ws = wb.create_sheet("Participants")
        
        # Headers
        headers = [
            'Email', 'Location', 'Country', 'Age', 'Gender', 'Gender Other',
            'Living With', 'Living With Other', 'University', 'Career',
            'Current Semester', 'Marital Status', 'Mother Education',
            'Father Education', 'Mother Age', 'Father Age', 'GPA Last Semester',
            'Repeated Cycles', 'Repeated Cycles Count', 'Residence Sector',
            'Socioeconomic Level', 'Income Sources', 'Consent Accepted',
            'Consent Date', 'Feedback Sent', 'Created At'
        ]
        
        self._write_header_row(ws, headers)
        
        # Data
        for participant in queryset:
            row = [
                participant.email,
                participant.get_location_display(),
                participant.country or '',
                participant.age or '',
                participant.get_gender_display() if participant.gender else '',
                participant.gender_other or '',
                participant.get_living_with_display() if participant.living_with else '',
                participant.living_with_other or '',
                participant.university or '',
                participant.career or '',
                participant.get_current_semester_display() if participant.current_semester else '',
                participant.get_marital_status_display() if participant.marital_status else '',
                participant.get_mother_education_level_display() if participant.mother_education_level else '',
                participant.get_father_education_level_display() if participant.father_education_level else '',
                participant.mother_age or '',
                participant.father_age or '',
                float(participant.gpa_last_semester) if participant.gpa_last_semester else '',
                'Sí' if participant.repeated_cycles else 'No',
                participant.repeated_cycles_count or '',
                participant.get_residence_sector_display() if participant.residence_sector else '',
                participant.get_socioeconomic_level_display() if participant.socioeconomic_level else '',
                participant.income_sources or '',
                'Sí' if participant.consent_accepted else 'No',
                participant.consent_date.strftime('%Y-%m-%d %H:%M:%S') if participant.consent_date else '',
                'Sí' if participant.feedback_sent else 'No',
                participant.created_at.strftime('%Y-%m-%d %H:%M:%S') if participant.created_at else '',
            ]
            ws.append(row)
        
        self._auto_adjust_column_width(ws)
    
    def _create_bergen_tiktok_sheet(self, wb, queryset):
        """Create sheet for Bergen TikTok data"""
        ws = wb.create_sheet("Bergen TikTok")
        
        headers = [
            'Email', 'Q1 Salience', 'Q2 Tolerance', 'Q3 Mood Modification',
            'Q4 Relapse', 'Q5 Withdrawal', 'Q6 Conflict', 'Total Score',
            'Feedback', 'Created At'
        ]
        
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'bergen_tiktok'):
                bt = participant.bergen_tiktok
                row = [
                    participant.email,
                    bt.q1_salience,
                    bt.q2_tolerance,
                    bt.q3_mood_modification,
                    bt.q4_relapse,
                    bt.q5_withdrawal,
                    bt.q6_conflict,
                    bt.total_score,
                    bt.get_feedback(),
                    bt.created_at.strftime('%Y-%m-%d %H:%M:%S') if bt.created_at else '',
                ]
                ws.append(row)
        
        self._auto_adjust_column_width(ws)
    
    def _create_bergen_instagram_sheet(self, wb, queryset):
        """Create sheet for Bergen Instagram data"""
        ws = wb.create_sheet("Bergen Instagram")
        
        headers = [
            'Email', 'Q1 Salience', 'Q2 Tolerance', 'Q3 Mood Modification',
            'Q4 Relapse', 'Q5 Withdrawal', 'Q6 Conflict', 'Total Score',
            'Feedback', 'Created At'
        ]
        
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'bergen_instagram'):
                bi = participant.bergen_instagram
                row = [
                    participant.email,
                    bi.q1_salience,
                    bi.q2_tolerance,
                    bi.q3_mood_modification,
                    bi.q4_relapse,
                    bi.q5_withdrawal,
                    bi.q6_conflict,
                    bi.total_score,
                    bi.get_feedback(),
                    bi.created_at.strftime('%Y-%m-%d %H:%M:%S') if bi.created_at else '',
                ]
                ws.append(row)
        
        self._auto_adjust_column_width(ws)
    
    def _create_ucla_loneliness_sheet(self, wb, queryset):
        """Create sheet for UCLA Loneliness data"""
        ws = wb.create_sheet("UCLA Loneliness")
        
        headers = ['Email'] + [f'Q{i}' for i in range(1, 21)] + ['Total Score', 'Feedback', 'Created At']
        
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'ucla_loneliness'):
                ul = participant.ucla_loneliness
                row = [participant.email]
                row.extend([getattr(ul, f'q{i}') for i in range(1, 21)])
                row.extend([
                    ul.total_score,
                    ul.get_feedback(),
                    ul.created_at.strftime('%Y-%m-%d %H:%M:%S') if ul.created_at else '',
                ])
                ws.append(row)
        
        self._auto_adjust_column_width(ws)
    
    def _create_prefrontal_symptoms_sheet(self, wb, queryset):
        """Create sheet for Prefrontal Symptoms data"""
        ws = wb.create_sheet("Prefrontal Symptoms")
        
        headers = ['Email'] + [f'Q{i}' for i in range(1, 21)] + ['Total Score', 'Feedback', 'Created At']
        
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'prefrontal_symptoms'):
                ps = participant.prefrontal_symptoms
                row = [participant.email]
                row.extend([getattr(ps, f'q{i}') for i in range(1, 21)])
                row.extend([
                    ps.total_score,
                    ps.get_feedback(),
                    ps.created_at.strftime('%Y-%m-%d %H:%M:%S') if ps.created_at else '',
                ])
                ws.append(row)
        
        self._auto_adjust_column_width(ws)
    
    def _create_caids_sheet(self, wb, queryset):
        """Create sheet for CAIDS data"""
        ws = wb.create_sheet("CAIDS")
        
        headers = ['Email'] + [f'Q{i}' for i in range(1, 14)] + ['Total Score', 'Feedback', 'Created At']
        
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'caids'):
                caids = participant.caids
                row = [participant.email]
                row.extend([getattr(caids, f'q{i}') for i in range(1, 14)])
                row.extend([
                    caids.total_score,
                    caids.get_feedback(),
                    caids.created_at.strftime('%Y-%m-%d %H:%M:%S') if caids.created_at else '',
                ])
                ws.append(row)
        
        self._auto_adjust_column_width(ws)
    
    def _create_summary_sheet(self, wb, queryset):
        """Create summary sheet with key statistics"""
        ws = wb.create_sheet("Summary", 0)  # Insert as first sheet
        
        # Title
        ws['A1'] = 'SURVEY DATA SUMMARY'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Export info
        ws['A3'] = 'Export Date:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = 'Total Participants:'
        ws['B4'] = queryset.count()
        
        # Statistics by instrument
        row = 6
        ws[f'A{row}'] = 'Instrument Statistics'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        instruments = [
            ('Bergen TikTok', 'bergen_tiktok'),
            ('Bergen Instagram', 'bergen_instagram'),
            ('UCLA Loneliness', 'ucla_loneliness'),
            ('Prefrontal Symptoms', 'prefrontal_symptoms'),
            ('CAIDS', 'caids'),
        ]
        
        for name, attr in instruments:
            count = sum(1 for p in queryset if hasattr(p, attr))
            ws[f'A{row}'] = name
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f'{(count/queryset.count()*100):.1f}%' if queryset.count() > 0 else '0%'
            row += 1
        
        # Demographics summary
        row += 2
        ws[f'A{row}'] = 'Demographics Summary'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Location breakdown
        for location in ['EC', 'CL']:
            count = queryset.filter(location=location).count()
            ws[f'A{row}'] = f'Location - {dict(Participant.LOCATION_CHOICES)[location]}'
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f'{(count/queryset.count()*100):.1f}%' if queryset.count() > 0 else '0%'
            row += 1
        
        # Gender breakdown
        for gender_code, gender_name in Participant.GENDER_CHOICES:
            count = queryset.filter(gender=gender_code).count()
            if count > 0:
                ws[f'A{row}'] = f'Gender - {gender_name}'
                ws[f'B{row}'] = count
                ws[f'C{row}'] = f'{(count/queryset.count()*100):.1f}%' if queryset.count() > 0 else '0%'
                row += 1
        
        self._auto_adjust_column_width(ws)
    
    def _write_header_row(self, ws, headers):
        """Write and style header row"""
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _auto_adjust_column_width(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


# Register models
admin.site.register(Participant, ParticipantAdmin)


@admin.register(BergenTikTok)
class BergenTikTokAdmin(admin.ModelAdmin):
    list_display = ['participant', 'total_score', 'created_at']
    search_fields = ['participant__email']


@admin.register(BergenInstagram)
class BergenInstagramAdmin(admin.ModelAdmin):
    list_display = ['participant', 'total_score', 'created_at']
    search_fields = ['participant__email']


@admin.register(UCLALoneliness)
class UCLALonelinessAdmin(admin.ModelAdmin):
    list_display = ['participant', 'total_score', 'created_at']
    search_fields = ['participant__email']


@admin.register(PrefrontalSymptoms)
class PrefrontalSymptomsAdmin(admin.ModelAdmin):
    list_display = ['participant', 'total_score', 'created_at']
    search_fields = ['participant__email']


@admin.register(CAIDS)
class CAIDSAdmin(admin.ModelAdmin):
    list_display = ['participant', 'total_score', 'created_at']
    search_fields = ['participant__email']