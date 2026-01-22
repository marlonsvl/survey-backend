# Save this file as: surveys/management/commands/export_survey_data.py

from django.core.management.base import BaseCommand
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from surveys.models import Participant
import os


class Command(BaseCommand):
    help = 'Export survey data to Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--location',
            type=str,
            help='Filter by location (EC or CL)',
        )
        parser.add_argument(
            '--start-date',
            type=str,
            help='Start date (YYYY-MM-DD)',
        )
        parser.add_argument(
            '--end-date',
            type=str,
            help='End date (YYYY-MM-DD)',
        )
        parser.add_argument(
            '--output',
            type=str,
            default='survey_export.xlsx',
            help='Output filename (default: survey_export.xlsx)',
        )
        parser.add_argument(
            '--output-dir',
            type=str,
            default='.',
            help='Output directory (default: current directory)',
        )

    def handle(self, *args, **options):
        # Build queryset
        queryset = Participant.objects.all()

        if options['location']:
            queryset = queryset.filter(location=options['location'])
            self.stdout.write(f"Filtering by location: {options['location']}")

        if options['start_date']:
            queryset = queryset.filter(created_at__gte=options['start_date'])
            self.stdout.write(f"Filtering from date: {options['start_date']}")

        if options['end_date']:
            queryset = queryset.filter(created_at__lte=options['end_date'])
            self.stdout.write(f"Filtering to date: {options['end_date']}")

        # Check if we have data
        count = queryset.count()
        if count == 0:
            self.stdout.write(self.style.WARNING('No data found with the given filters'))
            return

        self.stdout.write(f"Found {count} participants")

        # Create workbook
        self.stdout.write("Creating Excel workbook...")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Create sheets
        self._create_summary_sheet(wb, queryset)
        self._create_participants_sheet(wb, queryset)
        self._create_bergen_tiktok_sheet(wb, queryset)
        self._create_bergen_instagram_sheet(wb, queryset)
        self._create_ucla_loneliness_sheet(wb, queryset)
        self._create_prefrontal_symptoms_sheet(wb, queryset)
        self._create_caids_sheet(wb, queryset)

        # Save file
        output_path = os.path.join(options['output_dir'], options['output'])
        wb.save(output_path)

        self.stdout.write(
            self.style.SUCCESS(f'✅ Successfully exported data to: {output_path}')
        )
        self.stdout.write(f'📊 Total participants: {count}')
        
        # Show statistics
        instruments = [
            ('Bergen TikTok', 'bergen_tiktok'),
            ('Bergen Instagram', 'bergen_instagram'),
            ('UCLA Loneliness', 'ucla_loneliness'),
            ('Prefrontal Symptoms', 'prefrontal_symptoms'),
            ('CAIDS', 'caids'),
        ]
        
        self.stdout.write("\n📈 Instrument Completion:")
        for name, attr in instruments:
            inst_count = sum(1 for p in queryset if hasattr(p, attr))
            percentage = (inst_count / count * 100) if count > 0 else 0
            self.stdout.write(f"  - {name}: {inst_count}/{count} ({percentage:.1f}%)")

    def _create_summary_sheet(self, wb, queryset):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        ws['A1'] = 'SURVEY DATA EXPORT - COMMAND LINE'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = 'Export Date:'
        ws['B3'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = 'Total Records:'
        ws['B4'] = queryset.count()
        
        # Add filters info
        row = 6
        ws[f'A{row}'] = 'Applied Filters:'
        ws[f'A{row}'].font = Font(bold=True)
        
        self._auto_adjust_columns(ws)

    def _create_participants_sheet(self, wb, queryset):
        """Create participants sheet"""
        ws = wb.create_sheet("Participants")
        
        headers = [
            'Email', 'Location', 'Country', 'Age', 'Gender', 'Gender Other',
            'Living With', 'Living With Other', 'University', 'Career',
            'Current Semester', 'Marital Status', 'Mother Education',
            'Father Education', 'Mother Age', 'Father Age', 'GPA Last Semester',
            'Repeated Cycles', 'Repeated Cycles Count', 'Residence Sector',
            'Socioeconomic Level', 'Income Sources', 'Consent Accepted',
            'Consent Date', 'Feedback Sent', 'Created At'
        ]
        
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            row = [
                participant.email,
                participant.get_location_display(),
                participant.country or '',
                participant.age or '',
                participant.get_gender_display() if participant.gender else '',
                participant.gender_other or '',
                participant.get_living_with_display() if participant.living_with else '',
                participant.living_with_other or '',
                participant.university or '',
                participant.career or '',
                participant.get_current_semester_display() if participant.current_semester else '',
                participant.get_marital_status_display() if participant.marital_status else '',
                participant.get_mother_education_level_display() if participant.mother_education_level else '',
                participant.get_father_education_level_display() if participant.father_education_level else '',
                participant.mother_age or '',
                participant.father_age or '',
                float(participant.gpa_last_semester) if participant.gpa_last_semester else '',
                'Yes' if participant.repeated_cycles else 'No',
                participant.repeated_cycles_count or '',
                participant.get_residence_sector_display() if participant.residence_sector else '',
                participant.get_socioeconomic_level_display() if participant.socioeconomic_level else '',
                participant.income_sources or '',
                'Yes' if participant.consent_accepted else 'No',
                participant.consent_date.strftime('%Y-%m-%d %H:%M:%S') if participant.consent_date else '',
                'Yes' if participant.feedback_sent else 'No',
                participant.created_at.strftime('%Y-%m-%d %H:%M:%S') if participant.created_at else '',
            ]
            ws.append(row)
        
        self._auto_adjust_columns(ws)

    def _create_bergen_tiktok_sheet(self, wb, queryset):
        """Create Bergen TikTok sheet"""
        ws = wb.create_sheet("Bergen TikTok")
        
        headers = [
            'Email', 'Q1 Salience', 'Q2 Tolerance', 'Q3 Mood Modification',
            'Q4 Relapse', 'Q5 Withdrawal', 'Q6 Conflict', 'Total Score',
            'Feedback', 'Created At'
        ]
        
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'bergen_tiktok'):
                bt = participant.bergen_tiktok
                ws.append([
                    participant.email, bt.q1_salience, bt.q2_tolerance,
                    bt.q3_mood_modification, bt.q4_relapse, bt.q5_withdrawal,
                    bt.q6_conflict, bt.total_score, bt.get_feedback(),
                    bt.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
        
        self._auto_adjust_columns(ws)

    def _create_bergen_instagram_sheet(self, wb, queryset):
        """Create Bergen Instagram sheet"""
        ws = wb.create_sheet("Bergen Instagram")
        
        headers = [
            'Email', 'Q1 Salience', 'Q2 Tolerance', 'Q3 Mood Modification',
            'Q4 Relapse', 'Q5 Withdrawal', 'Q6 Conflict', 'Total Score',
            'Feedback', 'Created At'
        ]
        
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'bergen_instagram'):
                bi = participant.bergen_instagram
                ws.append([
                    participant.email, bi.q1_salience, bi.q2_tolerance,
                    bi.q3_mood_modification, bi.q4_relapse, bi.q5_withdrawal,
                    bi.q6_conflict, bi.total_score, bi.get_feedback(),
                    bi.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
        
        self._auto_adjust_columns(ws)

    def _create_ucla_loneliness_sheet(self, wb, queryset):
        """Create UCLA Loneliness sheet"""
        ws = wb.create_sheet("UCLA Loneliness")
        
        headers = ['Email'] + [f'Q{i}' for i in range(1, 21)] + ['Total Score', 'Feedback', 'Created At']
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'ucla_loneliness'):
                ul = participant.ucla_loneliness
                row = [participant.email]
                row.extend([getattr(ul, f'q{i}') for i in range(1, 21)])
                row.extend([
                    ul.total_score, ul.get_feedback(),
                    ul.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
                ws.append(row)
        
        self._auto_adjust_columns(ws)

    def _create_prefrontal_symptoms_sheet(self, wb, queryset):
        """Create Prefrontal Symptoms sheet"""
        ws = wb.create_sheet("Prefrontal Symptoms")
        
        headers = ['Email'] + [f'Q{i}' for i in range(1, 21)] + ['Total Score', 'Feedback', 'Created At']
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'prefrontal_symptoms'):
                ps = participant.prefrontal_symptoms
                row = [participant.email]
                row.extend([getattr(ps, f'q{i}') for i in range(1, 21)])
                row.extend([
                    ps.total_score, ps.get_feedback(),
                    ps.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
                ws.append(row)
        
        self._auto_adjust_columns(ws)

    def _create_caids_sheet(self, wb, queryset):
        """Create CAIDS sheet"""
        ws = wb.create_sheet("CAIDS")
        
        headers = ['Email'] + [f'Q{i}' for i in range(1, 14)] + ['Total Score', 'Feedback', 'Created At']
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'caids'):
                caids = participant.caids
                row = [participant.email]
                row.extend([getattr(caids, f'q{i}') for i in range(1, 14)])
                row.extend([
                    caids.total_score, caids.get_feedback(),
                    caids.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
                ws.append(row)
        
        self._auto_adjust_columns(ws)

    def _write_header_row(self, ws, headers):
        """Write and style header row"""
        ws.append(headers)
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width