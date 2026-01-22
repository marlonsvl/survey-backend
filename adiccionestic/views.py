from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from .models import (
    Participant, BergenTikTok, BergenInstagram,
    UCLALoneliness, PrefrontalSymptoms, CAIDS
)
from .serializers import (
    ParticipantSerializer, SurveySubmissionSerializer,
    BergenTikTokSerializer, BergenInstagramSerializer,
    UCLALonelinessSerializer, PrefrontalSymptomsSerializer,
    CAIDSSerializer
)

from django.views.generic import TemplateView

class ExportInterfaceView(TemplateView):
    template_name = 'export_interface.html'


class SurveyViewSet(viewsets.ModelViewSet):
    queryset = Participant.objects.all()
    serializer_class = ParticipantSerializer
    lookup_field = 'email'
    
    @action(detail=False, methods=['post'])
    def submit(self, request):
        """Submit survey responses"""
        serializer = SurveySubmissionSerializer(data=request.data)
        
        if serializer.is_valid():
            participant = serializer.save()
            
            # Send feedback email
            self.send_feedback_email(participant)
            
            return Response({
                'message': 'Encuesta enviada exitosamente',
                'email': participant.email
            }, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def feedback(self, request, email=None):
        """Get feedback for a participant"""
        try:
            participant = Participant.objects.get(email=email)
            feedback = self.generate_feedback(participant)
            return Response(feedback)
        except Participant.DoesNotExist:
            return Response(
                {'error': 'Participante no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export all survey data to Excel"""
        # Get filter parameters
        location = request.query_params.get('location', None)
        start_date = request.query_params.get('start_date', None)
        end_date = request.query_params.get('end_date', None)
        
        # Build queryset
        queryset = Participant.objects.all()
        
        if location:
            queryset = queryset.filter(location=location)
        
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)
        
        # Generate Excel file
        return self._generate_excel_export(queryset)

    @action(detail=True, methods=['get'])
    def export_participant(self, request, email=None):
        """Export single participant data to Excel"""
        try:
            participant = Participant.objects.get(email=email)
            queryset = Participant.objects.filter(email=email)
            return self._generate_excel_export(queryset)
        except Participant.DoesNotExist:
            return Response(
                {'error': 'Participante no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get summary statistics of survey data"""
        total_participants = Participant.objects.count()
        
        stats = {
            'total_participants': total_participants,
            'by_location': {},
            'by_gender': {},
            'instruments': {},
            'feedback_sent': Participant.objects.filter(feedback_sent=True).count(),
        }
        
        # Location stats
        for code, name in Participant.LOCATION_CHOICES:
            count = Participant.objects.filter(location=code).count()
            stats['by_location'][name] = {
                'count': count,
                'percentage': round(count / total_participants * 100, 2) if total_participants > 0 else 0
            }
        
        # Gender stats
        for code, name in Participant.GENDER_CHOICES:
            count = Participant.objects.filter(gender=code).count()
            if count > 0:
                stats['by_gender'][name] = {
                    'count': count,
                    'percentage': round(count / total_participants * 100, 2) if total_participants > 0 else 0
                }
        
        # Instrument completion stats
        instruments = [
            ('bergen_tiktok', 'Bergen TikTok'),
            ('bergen_instagram', 'Bergen Instagram'),
            ('ucla_loneliness', 'UCLA Loneliness'),
            ('prefrontal_symptoms', 'Prefrontal Symptoms'),
            ('caids', 'CAIDS'),
        ]
        
        for attr, name in instruments:
            count = sum(1 for p in Participant.objects.all() if hasattr(p, attr))
            stats['instruments'][name] = {
                'count': count,
                'percentage': round(count / total_participants * 100, 2) if total_participants > 0 else 0
            }
        
        return Response(stats)
    
    def _generate_excel_export(self, queryset):
        """Generate Excel file with survey data"""
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create sheets
        self._create_participants_sheet(wb, queryset)
        self._create_bergen_tiktok_sheet(wb, queryset)
        self._create_bergen_instagram_sheet(wb, queryset)
        self._create_ucla_loneliness_sheet(wb, queryset)
        self._create_prefrontal_symptoms_sheet(wb, queryset)
        self._create_caids_sheet(wb, queryset)
        self._create_summary_sheet(wb, queryset)
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename=survey_export_{timestamp}.xlsx'
        
        wb.save(response)
        return response
    
    def _create_participants_sheet(self, wb, queryset):
        """Create participants data sheet"""
        ws = wb.create_sheet("Participants")
        
        headers = [
            'Email', 'Location', 'Country', 'Age', 'Gender', 'Gender Other',
            'Living With', 'Living With Other', 'University', 'Career',
            'Current Semester', 'Marital Status', 'Mother Education',
            'Father Education', 'Mother Age', 'Father Age', 'GPA Last Semester',
            'Repeated Cycles', 'Repeated Cycles Count', 'Residence Sector',
            'Socioeconomic Level', 'Income Sources', 'Consent Accepted',
            'Consent Date', 'Feedback Sent', 'Created At'
        ]
        
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            row = [
                participant.email,
                participant.get_location_display(),
                participant.country or '',
                participant.age or '',
                participant.get_gender_display() if participant.gender else '',
                participant.gender_other or '',
                participant.get_living_with_display() if participant.living_with else '',
                participant.living_with_other or '',
                participant.university or '',
                participant.career or '',
                participant.get_current_semester_display() if participant.current_semester else '',
                participant.get_marital_status_display() if participant.marital_status else '',
                participant.get_mother_education_level_display() if participant.mother_education_level else '',
                participant.get_father_education_level_display() if participant.father_education_level else '',
                participant.mother_age or '',
                participant.father_age or '',
                float(participant.gpa_last_semester) if participant.gpa_last_semester else '',
                'Yes' if participant.repeated_cycles else 'No',
                participant.repeated_cycles_count or '',
                participant.get_residence_sector_display() if participant.residence_sector else '',
                participant.get_socioeconomic_level_display() if participant.socioeconomic_level else '',
                participant.income_sources or '',
                'Yes' if participant.consent_accepted else 'No',
                participant.consent_date.strftime('%Y-%m-%d %H:%M:%S') if participant.consent_date else '',
                'Yes' if participant.feedback_sent else 'No',
                participant.created_at.strftime('%Y-%m-%d %H:%M:%S') if participant.created_at else '',
            ]
            ws.append(row)
        
        self._auto_adjust_columns(ws)
    
    def _create_bergen_tiktok_sheet(self, wb, queryset):
        """Create Bergen TikTok sheet"""
        ws = wb.create_sheet("Bergen TikTok")
        
        headers = [
            'Email', 'Q1 Salience', 'Q2 Tolerance', 'Q3 Mood Modification',
            'Q4 Relapse', 'Q5 Withdrawal', 'Q6 Conflict', 'Total Score',
            'Feedback', 'Created At'
        ]
        
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'bergen_tiktok'):
                bt = participant.bergen_tiktok
                ws.append([
                    participant.email, bt.q1_salience, bt.q2_tolerance,
                    bt.q3_mood_modification, bt.q4_relapse, bt.q5_withdrawal,
                    bt.q6_conflict, bt.total_score, bt.get_feedback(),
                    bt.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
        
        self._auto_adjust_columns(ws)
    
    def _create_bergen_instagram_sheet(self, wb, queryset):
        """Create Bergen Instagram sheet"""
        ws = wb.create_sheet("Bergen Instagram")
        
        headers = [
            'Email', 'Q1 Salience', 'Q2 Tolerance', 'Q3 Mood Modification',
            'Q4 Relapse', 'Q5 Withdrawal', 'Q6 Conflict', 'Total Score',
            'Feedback', 'Created At'
        ]
        
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'bergen_instagram'):
                bi = participant.bergen_instagram
                ws.append([
                    participant.email, bi.q1_salience, bi.q2_tolerance,
                    bi.q3_mood_modification, bi.q4_relapse, bi.q5_withdrawal,
                    bi.q6_conflict, bi.total_score, bi.get_feedback(),
                    bi.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
        
        self._auto_adjust_columns(ws)
    
    def _create_ucla_loneliness_sheet(self, wb, queryset):
        """Create UCLA Loneliness sheet"""
        ws = wb.create_sheet("UCLA Loneliness")
        
        headers = ['Email'] + [f'Q{i}' for i in range(1, 21)] + ['Total Score', 'Feedback', 'Created At']
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'ucla_loneliness'):
                ul = participant.ucla_loneliness
                row = [participant.email]
                row.extend([getattr(ul, f'q{i}') for i in range(1, 21)])
                row.extend([
                    ul.total_score, ul.get_feedback(),
                    ul.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
                ws.append(row)
        
        self._auto_adjust_columns(ws)
    
    def _create_prefrontal_symptoms_sheet(self, wb, queryset):
        """Create Prefrontal Symptoms sheet"""
        ws = wb.create_sheet("Prefrontal Symptoms")
        
        headers = ['Email'] + [f'Q{i}' for i in range(1, 21)] + ['Total Score', 'Feedback', 'Created At']
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'prefrontal_symptoms'):
                ps = participant.prefrontal_symptoms
                row = [participant.email]
                row.extend([getattr(ps, f'q{i}') for i in range(1, 21)])
                row.extend([
                    ps.total_score, ps.get_feedback(),
                    ps.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
                ws.append(row)
        
        self._auto_adjust_columns(ws)
    
    def _create_caids_sheet(self, wb, queryset):
        """Create CAIDS sheet"""
        ws = wb.create_sheet("CAIDS")
        
        headers = ['Email'] + [f'Q{i}' for i in range(1, 14)] + ['Total Score', 'Feedback', 'Created At']
        self._write_header_row(ws, headers)
        
        for participant in queryset:
            if hasattr(participant, 'caids'):
                caids = participant.caids
                row = [participant.email]
                row.extend([getattr(caids, f'q{i}') for i in range(1, 14)])
                row.extend([
                    caids.total_score, caids.get_feedback(),
                    caids.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
                ws.append(row)
        
        self._auto_adjust_columns(ws)
    
    def _create_summary_sheet(self, wb, queryset):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        ws['A1'] = 'SURVEY DATA EXPORT'
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A3'] = 'Export Date:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = 'Total Records:'
        ws['B4'] = queryset.count()
        
        self._auto_adjust_columns(ws)
    
    def _write_header_row(self, ws, headers):
        """Write styled header row"""
        ws.append(headers)
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def generate_feedback(self, participant):
        """Generate feedback for all completed instruments"""
        feedback = {
            'email': participant.email,
            'instruments': {}
        }
        
        if hasattr(participant, 'bergen_tiktok'):
            feedback['instruments']['bergen_tiktok'] = {
                'score': participant.bergen_tiktok.total_score,
                'feedback': participant.bergen_tiktok.get_feedback()
            }
        
        if hasattr(participant, 'bergen_instagram'):
            feedback['instruments']['bergen_instagram'] = {
                'score': participant.bergen_instagram.total_score,
                'feedback': participant.bergen_instagram.get_feedback()
            }
        
        if hasattr(participant, 'ucla_loneliness'):
            feedback['instruments']['ucla_loneliness'] = {
                'score': participant.ucla_loneliness.total_score,
                'feedback': participant.ucla_loneliness.get_feedback()
            }
        
        if hasattr(participant, 'prefrontal_symptoms'):
            feedback['instruments']['prefrontal_symptoms'] = {
                'score': participant.prefrontal_symptoms.total_score,
                'feedback': participant.prefrontal_symptoms.get_feedback()
            }
        
        if hasattr(participant, 'caids'):
            feedback['instruments']['caids'] = {
                'score': participant.caids.total_score,
                'feedback': participant.caids.get_feedback()
            }
        
        return feedback
    
    def send_feedback_email(self, participant):
        """Send feedback email to participant using Mailgun"""
        from django.core.mail import EmailMultiAlternatives
        from django.template.loader import render_to_string
        
        feedback = self.generate_feedback(participant)
        
        subject = 'Resultados de tu Encuesta Psicológica'
        
        # Create email context
        context = {
            'email': participant.email,
            'location': participant.get_location_display(),
            'instruments': feedback['instruments'],
        }
        
        # Render HTML template
        html_message = self._render_feedback_html(context)
        
        # Create plain text version
        text_message = self._render_feedback_text(context)
        
        # Create email with both plain text and HTML
        email = EmailMultiAlternatives(
            subject=subject,
            body=text_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[participant.email],
        )
        
        # Attach HTML version
        email.attach_alternative(html_message, "text/html")
        
        try:
            # Send email via Mailgun
            result = email.send()
            
            if result:
                participant.feedback_sent = True
                participant.save()
                print(f"✅ Email sent successfully to {participant.email}")
            else:
                print(f"⚠️ Email sending returned False for {participant.email}")
                
        except Exception as e:
            print(f"❌ Error sending email to {participant.email}: {e}")
            raise

    def _render_feedback_html(self, context):
        """Render HTML email template"""
        from django.template.loader import render_to_string
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    line-height: 1.6;
                    color: #333;
                }}
                .container {{
                    max-width: 600px;
                    margin: 0 auto;
                    padding: 20px;
                    background-color: #f9f9f9;
                }}
                .header {{
                    background: linear-gradient(135deg, #6C63FF 0%, #8B7FFF 100%);
                    color: white;
                    padding: 30px;
                    border-radius: 10px 10px 0 0;
                    text-align: center;
                }}
                .header h1 {{
                    margin: 0;
                    font-size: 28px;
                    font-weight: bold;
                }}
                .content {{
                    background: white;
                    padding: 30px;
                }}
                .instrument {{
                    margin: 20px 0;
                    padding: 15px;
                    background: #f5f5f5;
                    border-left: 4px solid #6C63FF;
                    border-radius: 5px;
                }}
                .instrument-title {{
                    font-size: 16px;
                    font-weight: bold;
                    color: #6C63FF;
                    margin: 0 0 10px 0;
                }}
                .instrument-score {{
                    font-size: 14px;
                    color: #666;
                    margin: 5px 0;
                }}
                .instrument-feedback {{
                    font-size: 16px;
                    font-weight: bold;
                    color: #333;
                    margin: 10px 0 0 0;
                    padding: 10px;
                    background: white;
                    border-radius: 5px;
                }}
                .feedback-low {{
                    background-color: #d4edda;
                    color: #155724;
                }}
                .feedback-moderate {{
                    background-color: #fff3cd;
                    color: #856404;
                }}
                .feedback-high {{
                    background-color: #f8d7da;
                    color: #721c24;
                }}
                .footer {{
                    background: #f5f5f5;
                    padding: 20px;
                    border-radius: 0 0 10px 10px;
                    text-align: center;
                    font-size: 12px;
                    color: #666;
                }}
                .disclaimer {{
                    padding: 15px;
                    background: #e7f3ff;
                    border-left: 4px solid #2196F3;
                    margin: 20px 0;
                    font-size: 13px;
                    color: #0c5aa0;
                    border-radius: 5px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🎉 Encuesta Completada</h1>
                    <p>Aquí están tus resultados</p>
                </div>
                
                <div class="content">
                    <p>Hola,</p>
                    <p>Gracias por completar nuestra encuesta psicológica. A continuación encontrarás el resumen de tus resultados:</p>
                    
                    <div style="margin: 30px 0;">
        """
        
        # Add instrument results
        instrument_names = {
            'bergen_tiktok': '📱 Bergen TikTok - Escala de Adicción a TikTok',
            'bergen_instagram': '📸 Bergen Instagram - Escala de Adicción a Instagram',
            'ucla_loneliness': '👥 UCLA - Escala de Soledad',
            'prefrontal_symptoms': '🧠 Síntomas Prefrontales - Inventario Abreviado',
            'caids': '🤖 CAIDS - Dependencia de IA Conversacional',
        }
        
        for instrument, data in context['instruments'].items():
            feedback_class = self._get_feedback_class(data['feedback'])
            html += f"""
                    <div class="instrument">
                        <div class="instrument-title">{instrument_names.get(instrument, instrument)}</div>
                        <div class="instrument-score">Puntuación: {data['score']}</div>
                        <div class="instrument-feedback feedback-{feedback_class}">
                            {data['feedback']}
                        </div>
                    </div>
            """
        
        html += """
                    </div>
                    
                    <div class="disclaimer">
                        <strong>⚠️ Descargo de Responsabilidad:</strong><br>
                        Estos resultados son solo para fines informativos y no constituyen un diagnóstico clínico. 
                        Si tienes preocupaciones sobre tu salud mental o bienestar, te recomendamos consultar con un 
                        profesional de la salud mental calificado.
                    </div>
                    
                    <p>Si tienes preguntas sobre tus resultados o la encuesta, no dudes en contactarnos.</p>
                </div>
                
                <div class="footer">
                    <p>&copy; 2025 Estudio de Evaluación Psicológica. Todos los derechos reservados.</p>
                    <p>Este es un correo automatizado. Por favor no respondas a este mensaje.</p>
                </div>
            </div>
        </body>
        </html>
        """
        return html

    def _render_feedback_text(self, context):
        """Render plain text email"""
        text = "RESULTADOS DE TU ENCUESTA PSICOLÓGICA\n"
        text += "=" * 50 + "\n\n"
        text += f"Correo: {context['email']}\n"
        text += f"Ubicación: {context['location']}\n\n"
        
        instrument_names = {
            'bergen_tiktok': 'Bergen TikTok - Escala de Adicción a TikTok',
            'bergen_instagram': 'Bergen Instagram - Escala de Adicción a Instagram',
            'ucla_loneliness': 'UCLA - Escala de Soledad',
            'prefrontal_symptoms': 'Síntomas Prefrontales - Inventario Abreviado',
            'caids': 'CAIDS - Dependencia de IA Conversacional',
        }
        
        for instrument, data in context['instruments'].items():
            text += f"{instrument_names.get(instrument, instrument)}\n"
            text += "-" * 50 + "\n"
            text += f"Puntuación: {data['score']}\n"
            text += f"Resultado: {data['feedback']}\n\n"
        
        text += "=" * 50 + "\n"
        text += "DESCARGO DE RESPONSABILIDAD:\n"
        text += "Estos resultados son solo para fines informativos y no constituyen\n"
        text += "un diagnóstico clínico. Si tienes preocupaciones sobre tu salud mental,\n"
        text += "te recomendamos consultar con un profesional calificado.\n\n"
        text += "Equipo de Investigación"
        
        return text

    def _get_feedback_class(self, feedback_text):
        """Determine CSS class for feedback styling"""
        if 'bajo' in feedback_text.lower() or 'mínimo' in feedback_text.lower():
            return 'low'
        elif 'moderado' in feedback_text.lower() or 'leve' in feedback_text.lower():
            return 'moderate'
        else:
            return 'high'